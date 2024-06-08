# -*- encoding: UTF-8 -*-
from openpyxl import load_workbook
import time

# Mở tệp Excel
workbook = load_workbook('data_joint.xlsx')

# Chọn trang tính
sheet = workbook.active  # Trang tính mặc định

# Thông tin kết nối đến robot
from naoqi import ALProxy
robot_ip = "127.0.0.1"  # Thay bằng địa chỉ IP của robot NAO
robot_port = 9559
# Tạo proxy cho ALMotion
motion_proxy = ALProxy("ALMotion", robot_ip, robot_port)

for step in range(1,10):
    column_index = 6 + step
    print("Farme ",step)
    column_values = []

    # Lặp qua các hàng và lấy giá trị từ cột chỉ định
    for row in sheet.iter_rows(min_row=2):
        column_values.append(row[column_index].value)

    names  = "Body"
    angles  = map(float, column_values)
    fractionMaxSpeed  = 0.01
    motion_proxy.setAngles(names, angles, fractionMaxSpeed)
    time.sleep(10)