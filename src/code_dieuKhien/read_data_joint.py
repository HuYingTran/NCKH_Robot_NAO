# -*- encoding: UTF-8 -*-
from openpyxl import load_workbook

# Mở tệp Excel
workbook = load_workbook('data_joint.xlsx')

# Chọn trang tính
sheet = workbook.active  # Trang tính mặc định

# Lấy dữ liệu từ cột thứ n (index từ 1)
column_index = 1  # Đổi chỉ số này để lấy cột khác
column_values = []

# Lặp qua các hàng và lấy giá trị từ cột chỉ định
for row in sheet.iter_rows(min_row=2):
    column_values.append(row[column_index].value)

# In giá trị của cột
for value in column_values:
    print(value)
