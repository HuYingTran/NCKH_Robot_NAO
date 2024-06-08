# -*- encoding: UTF-8 -*-
from naoqi import ALProxy

# Thông tin kết nối đến robot
robot_ip = "127.0.0.1"  # Thay bằng địa chỉ IP của robot NAO
robot_port = 9559
# Tạo proxy cho ALMotion
motion_proxy = ALProxy("ALMotion", robot_ip, robot_port)
joint_data = []

from openpyxl import load_workbook

# Mở tệp Excel
workbook = load_workbook('data_joint.xlsx')
sheet = workbook.active  # Trang tính mặc định
column_index = len([cell for cell in sheet[2] if cell.value is not None]) + 1
index = 1

names         = "Body"
useSensors    = False
commandAngles = motion_proxy.getAngles(names, useSensors)
print "Command angles:"
print str(commandAngles)
print ""

sheet.cell(row=index, column = column_index, value = "Farme_" + str(column_index-6 ))
# Đọc giá trị góc các joint
for angles in commandAngles:
    index = index + 1
    sheet.cell(row=index, column = column_index, value = angles)

workbook.save('data_joint.xlsx')
print("Save Farme",column_index - 6)